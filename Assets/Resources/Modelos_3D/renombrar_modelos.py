from pathlib import Path
from openpyxl import load_workbook
import argparse
import unicodedata
import re


def limpiar_texto(texto):
    if texto is None:
        return ""
    texto = str(texto).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def normalizar_para_comparar(texto):
    texto = limpiar_texto(texto).lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"[^a-z0-9 ]", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def convertir_a_nombre_archivo(texto):
    texto = limpiar_texto(texto)
    texto = texto.replace("/", "-").replace("\\", "-")
    texto = re.sub(r"[^\w\sáéíóúÁÉÍÓÚñÑüÜ-]", "", texto)
    texto = re.sub(r"\s+", "_", texto)
    return texto


def extraer_estudiantes(celda):
    if celda is None:
        return []

    texto = str(celda).strip()

    if not texto:
        return []

    estudiantes = []

    # Separa por saltos de línea, comas o punto y coma
    partes = re.split(r"[\n,;]+", texto)

    for parte in partes:
        estudiante = limpiar_texto(parte)
        if estudiante:
            estudiantes.append(estudiante)

    return estudiantes


def cargar_mapa_estudiantes(excel_path, columnas_estudiantes):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    mapa = {}

    for row in ws.iter_rows(min_row=2):
        especie = limpiar_texto(row[0].value)

        if not especie:
            continue

        for col_index in columnas_estudiantes:
            celda = row[col_index - 1].value
            estudiantes = extraer_estudiantes(celda)

            for estudiante in estudiantes:
                clave = normalizar_para_comparar(estudiante)
                mapa[clave] = especie

    return mapa


def buscar_especie_para_carpeta(nombre_carpeta, mapa_estudiantes):
    clave_carpeta = normalizar_para_comparar(nombre_carpeta)

    if clave_carpeta in mapa_estudiantes:
        return mapa_estudiantes[clave_carpeta]

    # Búsqueda flexible por si hay pequeñas diferencias entre carpeta y Excel
    for estudiante_normalizado, especie in mapa_estudiantes.items():
        if estudiante_normalizado in clave_carpeta or clave_carpeta in estudiante_normalizado:
            return especie

    return None


def renombrar_archivos(directorio_raiz, excel_path, columnas_estudiantes, aplicar):
    directorio_raiz = Path(directorio_raiz)
    excel_path = Path(excel_path)

    mapa_estudiantes = cargar_mapa_estudiantes(excel_path, columnas_estudiantes)

    print(f"Estudiantes cargados desde el Excel: {len(mapa_estudiantes)}")
    print()

    cambios = []
    sin_especie = []
    omitidos = []

    for carpeta_estudiante in directorio_raiz.iterdir():
        if not carpeta_estudiante.is_dir():
            continue

        nombre_estudiante = carpeta_estudiante.name
        especie = buscar_especie_para_carpeta(nombre_estudiante, mapa_estudiantes)

        if especie is None:
            sin_especie.append(nombre_estudiante)
            continue

        especie_archivo = convertir_a_nombre_archivo(especie)

        for archivo in carpeta_estudiante.glob("*.obj"):
            if archivo.name.startswith(especie_archivo + "_"):
                omitidos.append(str(archivo))
                continue

            nuevo_nombre = f"{especie_archivo}_{archivo.name}"
            nuevo_path = archivo.with_name(nuevo_nombre)

            if nuevo_path.exists():
                print(f"ADVERTENCIA: ya existe el archivo destino, no se renombra:")
                print(f"  {nuevo_path}")
                continue

            cambios.append((archivo, nuevo_path))

    print("Cambios propuestos:")
    print()

    for origen, destino in cambios:
        print(f"{origen}")
        print(f"  -> {destino}")
        print()

    print(f"Total de archivos a renombrar: {len(cambios)}")
    print(f"Archivos omitidos porque ya tenían especie: {len(omitidos)}")
    print(f"Carpetas sin especie encontrada en el Excel: {len(sin_especie)}")
    print()

    if sin_especie:
        print("Carpetas sin coincidencia:")
        for nombre in sin_especie:
            print(f"  - {nombre}")
        print()

    if aplicar:
        for origen, destino in cambios:
            origen.rename(destino)

        print("Renombrado aplicado correctamente.")
    else:
        print("Simulación terminada. No se cambió ningún archivo.")
        print("Para aplicar los cambios, ejecuta de nuevo el comando agregando --aplicar")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()

    parser.add_argument("directorio_raiz")
    parser.add_argument("excel")

    parser.add_argument(
        "--columnas",
        nargs="+",
        type=int,
        default=[3],
        help="Columnas donde están los estudiantes. Por defecto: 3"
    )

    parser.add_argument(
        "--aplicar",
        action="store_true",
        help="Aplica el renombrado real"
    )

    args = parser.parse_args()

    renombrar_archivos(
        directorio_raiz=args.directorio_raiz,
        excel_path=args.excel,
        columnas_estudiantes=args.columnas,
        aplicar=args.aplicar
    )
